from django.contrib.auth.models import User, Group
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

import podrazdeleniya.models as pod
import users.models as users
from api.views import translit
from utils.common import get_system_name


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fixF = lambda s: s[:1].upper() + s[1:] if s else ''  # noqa: E731
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        fio = 0
        rgt = 1
        otd = 2
        login = 3
        otds = {}
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Сотрудник" in cells and "Должность" in cells and "Подразделение" in cells:
                    starts = True
                    fio = cells.index("Сотрудник")
                    rgt = cells.index("Должность")
                    otd = cells.index("Подразделение")
                    login = cells.index("Логин")
                    continue
            else:
                f = cells[fio].replace("   ", " ").replace("  ", " ").replace("  ", " ").strip().title()
                r = cells[rgt].replace(" ", "").replace(".", ",").split(",")
                o = fixF(cells[otd].replace("   ", " ").replace("  ", " ").replace("  ", " ").strip())
                account = cells[login]

                fs = f.split(" ")
                if len(fs) == 0:
                    continue
                self.stdout.write("-------------------")
                fso = [fs[0]]
                if len(fs) > 1:
                    fso.extend([x[0] for x in fs[1:]])

                if o not in otds:
                    ps = pod.Podrazdeleniya.objects.filter(title=o).first()
                    if not ps:
                        ps = pod.Podrazdeleniya(title=o)
                        self.stdout.write("Добавлено новое подразделение: " + o)
                        if "КДЛ" == o or "лаборатория" in o.lower():
                            ps.p_type = pod.Podrazdeleniya.LABORATORY
                        elif "отделение" in o.lower() or "консуль" in o.lower() or "кабинет" in o.lower() or "специалист" in o.lower() or "центр" in o.lower():
                            ps.p_type = pod.Podrazdeleniya.DEPARTMENT
                        else:
                            self.stdout.write(f"Необходимо настроить тип в {get_system_name()}")
                        ps.save()
                    otds[o] = ps
                o = otds[o]
                username = translit(''.join(fso).lower())

                if account and len(account) > 0:
                    if '@' in account:
                        username = account.split('@')[0]
                    else:
                        username = account

                us = users.DoctorProfile.objects.filter(fio=f, podrazdeleniye=o, user__username=username).first()
                if not us:
                    user = User.objects.create_user(username)
                    user.set_password("123456")
                    user.is_active = True
                    user.save()
                    us = users.DoctorProfile(user=user, fio=f, podrazdeleniye=o)
                    us.save()
                    us.get_fio_parts()
                    self.stdout.write("Добавлен пользователь " + username + ". Необходимо сменить пароль (по умолчанию 123456)!")
                user = us.user
                user.groups.clear()
                for rpk in r:
                    g = Group.objects.get(pk=rpk)
                    user.groups.add(g)
